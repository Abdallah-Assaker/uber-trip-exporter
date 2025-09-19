"""
Uber Trip Exporter Script

This script extracts Uber trip data and fills out the monthly transportation claim form.

IMPORTANT: Before running this script:
1. Update the 'token.txt' file with your current Uber authentication cookie
2. Get the cookie from your browser's Developer Tools (F12) -> Network tab -> Cookie header
3. Do NOT put quotes around the cookie in the token.txt file
4. Configure your address keywords in the 'config.json' file (will be created automatically on first run)

Usage:
    python uber-script.py [month]
    
    month: Optional integer (1-12) for the month to fetch data for.
           If not provided, uses the previous month.
           For December (12), uses the previous year.

The script will:
- Fetch trip data from Uber's GraphQL API for the specified month
- Download receipt PDFs
- Fill out the Excel claim form using keyword-based address matching
- Merge all receipts into one PDF

For detailed instructions, see README.md
"""

# Standard library imports
import json
import os
import re
import shutil
import sys
import zipfile
from calendar import monthrange
from datetime import datetime, timedelta
from pathlib import Path

# Third-party imports
import pandas as pd
import requests
import time
from openpyxl import load_workbook
from PyPDF2 import PdfMerger

try:
    from redmail import EmailSender
except ImportError:
    EmailSender = None

# ============================================================================
# CONSTANTS AND CONFIGURATION
# ============================================================================

# API Configuration
UBER_GRAPHQL_URL = "https://riders.uber.com/graphql"

# Console colors for better logging
class Colors:
    HEADER = '\033[95m'
    BLUE = '\033[94m'
    CYAN = '\033[96m'
    GREEN = '\033[92m'
    WARNING = '\033[93m'
    ERROR = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'

def log(message, level="INFO"):
    """Enhanced console logging with colors and timestamps"""
    timestamp = datetime.now().strftime("%H:%M:%S")
    
    if level == "INFO":
        color = Colors.CYAN
    elif level == "SUCCESS":
        color = Colors.GREEN
    elif level == "WARNING":
        color = Colors.WARNING
    elif level == "ERROR":
        color = Colors.ERROR
    elif level == "HEADER":
        color = Colors.HEADER + Colors.BOLD
    else:
        color = Colors.ENDC
    
    print(f"{color}[{timestamp}] {level}: {message}{Colors.ENDC}")

def log_progress(current, total, message="Processing"):
    """Show progress with percentage"""
    percentage = (current / total) * 100 if total > 0 else 0
    print(f"{Colors.BLUE}[{datetime.now().strftime('%H:%M:%S')}] PROGRESS: {message} [{current}/{total}] ({percentage:.1f}%){Colors.ENDC}")

# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def get_month_date_range(month=None):
    """
    Calculate start and end timestamps for a given month.
    
    Args:
        month (int): Month number (1-12). If None, uses previous month.
                    For December (12), uses previous year.
    
    Returns:
        tuple: (start_timestamp_ms, end_timestamp_ms, month_year_string)
    """
    log("Calculating date range for requested month", "INFO")
    current_date = datetime.now()
    
    if month is None:
        # Default to previous month
        if current_date.month == 1:
            target_month = 12
            target_year = current_date.year - 1
        else:
            target_month = current_date.month - 1
            target_year = current_date.year
        log(f"No month specified, using previous month: {target_month}/{target_year}", "INFO")
    else:
        # Use provided month
        if not 1 <= month <= 12:
            log(f"Invalid month: {month}. Month must be between 1 and 12", "ERROR")
            raise ValueError("Month must be between 1 and 12")
        
        target_month = month
        if month == 12:
            target_year = current_date.year - 1
        else:
            target_year = current_date.year
        log(f"Using specified month: {target_month}/{target_year}", "INFO")
    
    # Calculate start of month (00:00:00)
    start_date = datetime(target_year, target_month, 1)
    
    # Calculate end of month (23:59:59.999)
    _, last_day = monthrange(target_year, target_month)
    end_date = datetime(target_year, target_month, last_day, 23, 59, 59, 999000)
    
    # Convert to milliseconds (Uber API expects milliseconds)
    start_timestamp_ms = int(start_date.timestamp() * 1000)
    end_timestamp_ms = int(end_date.timestamp() * 1000)
    
    # Create month-year string for file naming
    month_year_string = start_date.strftime("%Y-%m")
    
    log(f"Date range calculated: {start_date.strftime('%B %Y')}", "SUCCESS")
    log(f"Start: {start_date.strftime('%Y-%m-%d %H:%M:%S')}", "INFO")
    log(f"End: {end_date.strftime('%Y-%m-%d %H:%M:%S')}", "INFO")
    
    return start_timestamp_ms, end_timestamp_ms, month_year_string

def parse_command_line_args():
    """Parse command line arguments for month parameter."""
    log("Parsing command line arguments", "INFO")
    
    if len(sys.argv) > 1:
        try:
            month = int(sys.argv[1])
            if not 1 <= month <= 12:
                log("Month must be between 1 and 12", "ERROR")
                sys.exit(1)
            log(f"Month parameter provided: {month}", "SUCCESS")
            return month
        except ValueError:
            log("Invalid month parameter. Must be an integer between 1 and 12", "ERROR")
            log("Usage: python uber-script.py [month]", "INFO")
            log("Example: python uber-script.py 7  (for July)", "INFO")
            sys.exit(1)
    else:
        log("No month parameter provided, will use previous month", "INFO")
        return None

# ============================================================================
# CONFIGURATION AND DATA LOADING FUNCTIONS
# ============================================================================

def read_token_from_file(file_path="token.txt"):
    """
    Read the authentication token/cookie from a text file.
    The file should contain the cookie string without quotes.
    """
    log(f"Reading authentication token from {file_path}", "INFO")
    
    try:
        # Get the directory where the script is located
        script_dir = os.path.dirname(os.path.abspath(__file__))
        token_file_path = os.path.join(script_dir, file_path)
        
        with open(token_file_path, 'r', encoding='utf-8') as file:
            token = file.read().strip()
            if not token:
                raise ValueError("Token file is empty")
            log("Authentication token loaded successfully", "SUCCESS")
            return token
    except FileNotFoundError:
        log(f"{file_path} not found in the script directory", "ERROR")
        log(f"Please create a '{file_path}' file with your authentication cookie", "ERROR")
        log("You can copy the cookie from your browser's developer tools", "INFO")
        exit(1)
    except Exception as e:
        log(f"Error reading token file: {e}", "ERROR")
        exit(1)

def read_config_from_file(file_path="config.json"):
    """
    Read address keywords and email configuration from a JSON file.
    If the file doesn't exist, create it with default values.
    """
    log(f"Loading configuration from {file_path}", "INFO")
    
    try:
        # Get the directory where the script is located
        script_dir = os.path.dirname(os.path.abspath(__file__))
        config_file_path = os.path.join(script_dir, file_path)
        
        with open(config_file_path, 'r', encoding='utf-8') as file:
            config = json.load(file)
            
        # Validate required keys
        if 'home_address_keywords' not in config or 'work_address_keywords' not in config:
            raise ValueError("Config file must contain 'home_address_keywords' and 'work_address_keywords'")
        
        # Check for email configuration (optional)
        email_config = config.get('email_config', {})
        
        log(f"Configuration loaded successfully from {file_path}", "SUCCESS")
        log(f"Home keywords: {len(config['home_address_keywords'])} items", "INFO")
        log(f"Work keywords: {len(config['work_address_keywords'])} items", "INFO")
        
        if email_config:
            email_enabled = email_config.get('enabled', False)
            log(f"Email functionality: {'enabled' if email_enabled else 'disabled'}", "INFO")
        else:
            log("Email configuration not found - email functionality disabled", "WARNING")
            
        return config['home_address_keywords'], config['work_address_keywords'], email_config
        
    except FileNotFoundError:
        log(f"Config file {file_path} not found. Creating default config...", "WARNING")
        
        # Create default configuration
        default_config = {
            "home_address_keywords": [
                "AZ Zaytoun Al Qebleyah",
                "Zeitoun",
                "4520101"
            ],
            "work_address_keywords": [
                "223 متفرع من شارع 90",
                "224 N Teseen St",
                "4730420",
                "11835"
            ],
            "email_config": {
                "enabled": False,
                "recipient_email": "your-work-email@company.com",
                "sender_email": "your-sender-email@gmail.com",
                "sender_password": "your-app-password",
                "smtp_server": "smtp.gmail.com",
                "smtp_port": 587,
                "subject_template": "Uber Trip Report - {month_year}",
                "body_template": "Please find attached the Uber trip report for {month_year}.\n\nTotal amount: ${total_amount}\nNumber of trips: {trip_count}\n\nBest regards,\nUber Trip Exporter"
            },
            "_instructions": {
                "description": "Update the keyword lists above with parts of your addresses that are consistent",
                "email_setup": {
                    "description": "Configure email settings to automatically send trip reports",
                    "steps": [
                        "1. Set 'enabled' to true to enable email functionality",
                        "2. Update 'recipient_email' with your work email address",
                        "3. For Gmail: use your Gmail address as 'sender_email'",
                        "4. For Gmail: generate an App Password (not your regular password)",
                        "5. Update 'sender_password' with the App Password",
                        "6. For other email providers: update smtp_server and smtp_port accordingly"
                    ],
                    "gmail_app_password_guide": "https://support.google.com/accounts/answer/185833"
                },
                "tips": [
                    "Use street names, landmarks, or area names that appear in your addresses",
                    "Include both Arabic and English variations if applicable",
                    "Add multiple keywords for each location to handle address variations",
                    "Test by running the script and checking the trips.json file for actual addresses"
                ],
                "example": {
                    "home_address_keywords": [
                        "223 متفرع من شارع 90",
                        "خلف فندق الدوسيت", 
                        "N Teseen, New Cairo 1"
                    ],
                    "work_address_keywords": [
                        "1 Al Tabeer",
                        "El-Zaytoun Sharkeya",
                        "Zeitoun, Cairo"
                    ]
                }
            }
        }
        
        # Save default config
        with open(config_file_path, 'w', encoding='utf-8') as file:
            json.dump(default_config, file, ensure_ascii=False, indent=2)
            
        log(f"Created default config file: {file_path}", "SUCCESS")
        log(f"Please update the keywords in {file_path} with your actual address keywords", "WARNING")
        log("Email functionality is disabled by default. Configure email settings if needed.", "INFO")
        log("Script will exit. Please configure your addresses and run again.", "ERROR")
        exit(1)
        
    except json.JSONDecodeError as e:
        log(f"Invalid JSON format in {file_path}", "ERROR")
        log(f"JSON Error: {e}", "ERROR")
        exit(1)
    except Exception as e:
        log(f"Error reading config file: {e}", "ERROR")
        exit(1)

# ============================================================================
# API FUNCTIONS
# ============================================================================

def get_uber_trips(cookie, start_time_ms, end_time_ms, download_receipts=True):
    """
    Fetch Uber trips from GraphQL API for the specified time range.
    
    Args:
        cookie (str): Authentication cookie
        start_time_ms (int): Start timestamp in milliseconds
        end_time_ms (int): End timestamp in milliseconds
    
    Returns:
        tuple: (trips, overall_amount)
    """
    log("Fetching trips from Uber API...", "INFO")
    
    url = UBER_GRAPHQL_URL
    headers = {
        "cookie": cookie,
        "content-type": "application/json",
        "Cache-Control": "no-cache",
        "User-Agent": "PostmanRuntime/7.45.0",
        "origin": "https://riders.uber.com",
        "x-csrf-token": "x",
    }

    # 1. Query to list trips
    activities_query = """
query Activities(
  $cityID: Int
  $endTimeMs: Float
  $includePast: Boolean = true
  $includeUpcoming: Boolean = true
  $limit: Int = 60
  $nextPageToken: String
  $orderTypes: [RVWebCommonActivityOrderType!] = [RIDES, TRAVEL]
  $profileType: RVWebCommonActivityProfileType = PERSONAL
  $startTimeMs: Float
) {
  activities(cityID: $cityID) {
    past(
      endTimeMs: $endTimeMs
      limit: $limit
      nextPageToken: $nextPageToken
      orderTypes: $orderTypes
      profileType: $profileType
      startTimeMs: $startTimeMs
    ) @include(if: $includePast) {
      activities {
        uuid
        cardURL
        description
        subtitle  
        __typename
      }
      nextPageToken
      __typename
    }
    upcoming @include(if: $includeUpcoming) {
      activities {
        uuid
        cardURL
        description
        subtitle
        __typename
      }
      __typename
    }
    __typename
  }
}
"""

    # 2. Query to get details of a single trip (pickup & dropoff)
    trip_details_query = """
query GetTrip($tripUUID: String!) {
  getTrip(tripUUID: $tripUUID) {
    trip {
      uuid
      waypoints
    }
  }
}
"""

    # 3. Receipt query
    receipt_query = """
query GetReceipt($tripUUID: String!, $timestamp: String) {
  getReceipt(tripUUID: $tripUUID, timestamp: $timestamp) {
    receiptsForJob {
      timestamp
      type
    }
    receiptData
  }
}
"""

    # Variables for the activities query
    variables = {
        "endTimeMs": end_time_ms,
        "startTimeMs": start_time_ms,
        "limit": 60,
        "includePast": True,
        "includeUpcoming": False,
        "orderTypes": ["RIDES", "TRAVEL"],
        "profileType": "PERSONAL"
    }

    payload = {
        "operationName": "Activities",
        "query": activities_query,
        "variables": variables,
    }

    log("Making API request to fetch trip list...", "INFO")
    try:
        response = requests.post(url, headers=headers, data=json.dumps(payload), timeout=30)
    except requests.exceptions.Timeout:
        log("Timeout while fetching trips from API", "ERROR")
        return [], 0.0
    except requests.exceptions.RequestException as e:
        log(f"Network error while fetching trips: {e}", "ERROR")
        return [], 0.0

    if response.status_code != 200:
        log(f"Failed to fetch trips: HTTP {response.status_code}", "ERROR")
        return [], 0.0

    data = response.json()
    
    if not data or "data" not in data or not data["data"]:
        log("No data returned from API", "ERROR")
        return [], 0.0
        
    activities_data = data["data"].get("activities", {}).get("past", {}).get("activities", [])
    log(f"Found {len(activities_data)} trips in API response", "INFO")

    trips = []
    overall_amount = 0.0

    for i, trip in enumerate(activities_data):
        log_progress(i + 1, len(activities_data), "Processing trips")
        
        uuid = trip["uuid"]
        trip_url = trip["cardURL"]
        desc = trip.get("description", "")   
        subtitle = trip.get("subtitle", "")  

        # Parse price from description
        match = re.search(r"([0-9]+(?:\.[0-9]+)?)", desc)
        price = float(match.group(1)) if match else 0.0
        overall_amount += price

        # Check if trip was canceled
        status = "Canceled" if "canceled" in desc.lower() else "Completed"
        if status == "Canceled" or "unfulfilled" in desc.lower():
            log(f"Skipping canceled trip: {uuid}", "WARNING")
            continue

        pickup_address = ""
        dropoff_address = ""

        # Fetch trip details (pickup & dropoff addresses)
        trip_payload = {
            "operationName": "GetTrip",
            "query": trip_details_query,
            "variables": {"tripUUID": uuid},
        }

        try:
            detail_resp = requests.post(url, headers=headers, data=json.dumps(trip_payload), timeout=15)
        except requests.exceptions.Timeout:
            log(f"Timeout fetching trip details for {uuid}", "WARNING")
            detail_resp = None
        except requests.exceptions.RequestException as e:
            log(f"Network error fetching trip details for {uuid}: {e}", "WARNING")
            detail_resp = None
        
        if detail_resp and detail_resp.status_code == 200:
            detail_data = detail_resp.json()
            if detail_data and "data" in detail_data and detail_data["data"]:
                trip_info = detail_data["data"].get("getTrip", {})
                if trip_info and "trip" in trip_info:
                    trip_data = trip_info["trip"]
                    waypoints = trip_data.get("waypoints", [])
                    
                    if len(waypoints) >= 2:
                        # Waypoints are directly strings, not objects with 'name' property
                        pickup_address = waypoints[0] if isinstance(waypoints[0], str) else "Unknown pickup"
                        dropoff_address = waypoints[-1] if isinstance(waypoints[-1], str) else "Unknown dropoff"
                    else:
                        log(f"Insufficient waypoint data for trip {uuid}", "WARNING")
                else:
                    log(f"No trip data found for {uuid}", "WARNING")
            else:
                log(f"No valid data in response for {uuid}", "WARNING")
        else:
            log(f"Failed to fetch trip details for {uuid}", "WARNING")

        # Get receipt timestamp and download PDF (if enabled)
        if download_receipts:
            timestamp = get_receipt_timestamp(uuid, headers)
            if timestamp:
                download_receipt_pdf(uuid, timestamp, headers)
            else:
                log(f"No receipt timestamp found for trip {uuid}", "WARNING")
        else:
            log(f"Skipping receipt download for trip {uuid} (receipts disabled)", "INFO")

        trips.append({
            "uuid": uuid,
            "url": trip_url,
            "status": status,
            "price": price,
            "time": subtitle,
            "pickup_location": pickup_address,
            "dropoff_location": dropoff_address,
        })
        
        # Add delay between API calls to avoid rate limiting
        time.sleep(0.5)

    log(f"Successfully processed {len(trips)} trips", "SUCCESS")
    log(f"Total amount: ${overall_amount:.2f}", "INFO")
    
    return trips, overall_amount

# ============================================================================
# PDF AND RECEIPT FUNCTIONS
# ============================================================================

def download_receipt_pdf(uuid, timestamp, headers, folder="receipts", max_retries=3, timeout=30):
    """Download receipt PDF for a specific trip with retry logic."""
    os.makedirs(folder, exist_ok=True)
    pdf_path = os.path.join(folder, f"{uuid}.pdf")

    url = f"https://riders.uber.com/trips/{uuid}/receipt?contentType=PDF&timestamp={timestamp}"

    log(f"Downloading receipt for trip {uuid}", "INFO")
    
    for attempt in range(max_retries):
        try:
            resp = requests.get(url, headers=headers, timeout=timeout)
            if resp.status_code == 200 and resp.headers.get("content-type", "").startswith("application/pdf"):
                with open(pdf_path, "wb") as f:
                    f.write(resp.content)
                log(f"Successfully saved receipt: {pdf_path}", "SUCCESS")
                return pdf_path
            else:
                log(f"Failed to download receipt for {uuid}: HTTP {resp.status_code}", "WARNING")
                if attempt < max_retries - 1:
                    log(f"Retrying download for {uuid} (attempt {attempt + 2}/{max_retries})", "INFO")
                    time.sleep(2)  # Wait before retry
                continue
        except requests.exceptions.Timeout:
            log(f"Timeout downloading receipt for {uuid} (attempt {attempt + 1}/{max_retries})", "WARNING")
            if attempt < max_retries - 1:
                time.sleep(2)
        except requests.exceptions.RequestException as e:
            log(f"Network error downloading receipt for {uuid}: {e}", "WARNING")
            if attempt < max_retries - 1:
                time.sleep(2)
        except Exception as e:
            log(f"Unexpected error downloading receipt for {uuid}: {e}", "ERROR")
            break
    
    log(f"Failed to download receipt for {uuid} after {max_retries} attempts", "ERROR")
    return None

def cleanup_temp_receipts_folder(folder="receipts"):
    """Remove the temporary receipts folder after processing."""
    if os.path.exists(folder):
        shutil.rmtree(folder)
        log(f"Cleaned up temporary folder: {folder}", "INFO")
    else:
        log(f"Temporary folder {folder} not found, skipping cleanup", "INFO")
        print(f"🗑️ Cleaned up temporary folder: {folder}")

def create_monthly_excel_copy(template_file, month_year=None, output_folder=None):
    """
    Create a copy of the Excel template with month-year prefix in the specified output folder.
    If month_year is None, it will use the current month-year.
    If output_folder is None, it will create the file in the current directory.
    """
    log(f"Creating monthly Excel copy from template: {template_file}", "INFO")
    
    if month_year is None:
        month_year = datetime.now().strftime("%Y-%m")
    
    # Extract file name and extension
    file_name, file_ext = os.path.splitext(template_file)
    
    # Create new filename with month prefix
    new_filename = f"{month_year}_{file_name}{file_ext}"
    
    # If output folder is specified, create the path within that folder
    if output_folder:
        new_filepath = os.path.join(output_folder, new_filename)
    else:
        new_filepath = new_filename
    
    # Copy the template to new file
    shutil.copy2(template_file, new_filepath)
    log(f"Created monthly copy: {new_filepath}", "SUCCESS")
    
    return new_filepath

def process_excel_file(excel_file, trips, template_excel_file, home_keywords, work_keywords):
    """Process the Excel file with trip data."""
    log(f"Processing Excel file: {excel_file}", "INFO")
    log(f"Processing {len(trips)} trips", "INFO")
    
    def classify_trip_reason(pickup_location, home_keywords, work_keywords):
        """Classify trip reason based on pickup location keywords."""
        pickup_lower = pickup_location.lower()
        
        # Check if any home keyword is in the pickup location
        for keyword in home_keywords:
            if keyword.lower() in pickup_lower:
                return "العودة من العمل"  # Return from work
        
        # Check if any work keyword is in the pickup location  
        for keyword in work_keywords:
            if keyword.lower() in pickup_lower:
                return "الذهاب إلى العمل"  # Going to work
                
        return ""  # Unknown reason
    
    # Open the workbook (use the monthly copy we created)
    wb = load_workbook(excel_file)
    ws = wb["Claim Form"]

    start_row = 8  # Row 8 in Excel (since Excel rows are 1-based)

    for i, trip in enumerate(trips, start=0):
        log_progress(i + 1, len(trips), "Processing Excel rows")
        
        # Try to parse the trip date
        try:
            # Primary format: "Aug 31 • 4:29 PM"
            date_clean = trip["time"].replace("•", "").strip()
            trip_dt = datetime.strptime(date_clean, "%b %d %I:%M %p")
            trip_dt = trip_dt.replace(year=datetime.now().year)
            trip_date = trip_dt
        except:
            try:
                # Alternative format: "Aug 31, 2025, 4:29 PM"
                trip_dt = datetime.strptime(
                    trip["time"].replace("•", "").strip(), "%b %d, %Y, %I:%M %p"
                )
                trip_date = trip_dt
            except:
                trip_date = trip["time"]  # fallback if parsing fails

        # Classify trip reason for Excel column
        trip_reason_excel = classify_trip_reason(trip["pickup_location"], home_keywords, work_keywords)

        row = start_row + i
        ws.cell(row=row, column=2, value=trip_date)  # Column B = Date
        ws.cell(row=row, column=3, value=trip["pickup_location"])
        ws.cell(row=row, column=4, value=trip["dropoff_location"])
        ws.cell(row=row, column=5, value=trip["price"])
        ws.cell(row=row, column=6, value=trip_reason_excel)
        ws.cell(row=row, column=7, value="App Wallet")
        ws.cell(row=row, column=8, value="")

        # Format the date column as dd/mm/yyyy
        ws.cell(row=row, column=2).number_format = "dd/mm/yyyy"

    # Save changes to the monthly copy
    wb.save(excel_file)
    log(f"Data written to monthly copy: {excel_file}", "SUCCESS")
    log(f"Original template preserved: {template_excel_file}", "INFO")
  
def get_receipt_timestamp(uuid, headers):
    """Get the timestamp for a trip receipt."""
    url = UBER_GRAPHQL_URL
    
    receipt_query = """
query GetReceipt($tripUUID: String!, $timestamp: String) {
  getReceipt(tripUUID: $tripUUID, timestamp: $timestamp) {
    receiptsForJob {
      timestamp
      type
    }
    receiptData
  }
}
"""
    
    receipt_payload = {
        "operationName": "GetReceipt",
        "query": receipt_query,
        "variables": {"tripUUID": uuid, "timestamp": ""},
    }
    
    log(f"Getting receipt timestamp for trip {uuid}", "INFO")
    try:
        receipt_resp = requests.post(url, headers=headers, data=json.dumps(receipt_payload), timeout=15)
    except requests.exceptions.Timeout:
        log(f"Timeout getting receipt timestamp for {uuid}", "WARNING")
        return None
    except requests.exceptions.RequestException as e:
        log(f"Network error getting receipt timestamp for {uuid}: {e}", "WARNING")
        return None
        
    if receipt_resp.status_code == 200:
        receipt_data = receipt_resp.json()

        # Defensive checks
        if not receipt_data or "data" not in receipt_data or not receipt_data["data"]:
            log(f"No data returned for receipt {uuid}", "WARNING")
            return None

        receipt_info = receipt_data["data"].get("getReceipt")
        if receipt_info:
            jobs = receipt_info.get("receiptsForJob", [])
            if jobs:
                log(f"Successfully retrieved timestamp for trip {uuid}", "SUCCESS")
                return jobs[0]["timestamp"]
    else:
        log(f"Request failed for {uuid}: {receipt_resp.status_code}", "ERROR")
    return None

def merge_receipts(trips, folder="receipts", output_file="all_receipts.pdf"):
    """Merge all trip receipt PDFs into a single file."""
    log(f"Merging {len(trips)} receipts into {output_file}", "INFO")
    
    merger = PdfMerger()

    # Sort trips by time desc
    sorted_trips = sorted(
        trips,
        key=lambda t: parse_trip_date(t.get("time", "")),
        reverse=True
    )

    merged_count = 0
    for trip in sorted_trips:
        pdf_file = os.path.join(folder, f"{trip['uuid']}.pdf")
        if os.path.exists(pdf_file):
            merger.append(pdf_file)
            log(f"Added receipt for {trip['time']}", "INFO")
            merged_count += 1
        else:
            log(f"Missing PDF for {trip['uuid']} ({trip['time']})", "WARNING")

    if merged_count > 0:
        with open(output_file, "wb") as f:
            merger.write(f)
        merger.close()
        log(f"Successfully merged {merged_count} receipts into {output_file}", "SUCCESS")
    else:
        log("No trips available to merge", "WARNING")


def parse_trip_date(date_str: str):
    """Convert Uber's subtitle string into datetime (adjust format if needed)."""
    try:
        # Primary format: "Aug 31 • 4:29 PM"
        date_clean = date_str.replace("•", "").strip()
        trip_dt = datetime.strptime(date_clean, "%b %d %I:%M %p")
        trip_dt = trip_dt.replace(year=datetime.now().year)
        return trip_dt
    except Exception:
        try:
            # Alternative format: "Aug 31, 2025, 10:15 AM"
            return datetime.strptime(date_str, "%b %d, %Y, %I:%M %p")
        except Exception as e:
            log(f"Failed to parse date '{date_str}': {e}", "WARNING")
            return datetime.min    

# ============================================================================
# EMAIL AND ZIP FUNCTIONS
# ============================================================================

def create_zip_archive(source_folder, zip_filename):
    """
    Create a ZIP archive of the specified folder.
    
    Args:
        source_folder (str): Path to the folder to compress
        zip_filename (str): Name of the output ZIP file
    
    Returns:
        str: Path to the created ZIP file or None if failed
    """
    log(f"Creating ZIP archive: {zip_filename}", "INFO")
    
    try:
        with zipfile.ZipFile(zip_filename, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Walk through all files in the source folder
            for root, dirs, files in os.walk(source_folder):
                for file in files:
                    file_path = os.path.join(root, file)
                    # Calculate the relative path for the file in the ZIP
                    arcname = os.path.relpath(file_path, os.path.dirname(source_folder))
                    zip_file.write(file_path, arcname)
                    log(f"Added to ZIP: {arcname}", "INFO")
        
        # Check if ZIP file was created and get its size
        if os.path.exists(zip_filename):
            zip_size = os.path.getsize(zip_filename)
            log(f"ZIP archive created successfully: {zip_filename} ({zip_size / (1024*1024):.1f} MB)", "SUCCESS")
            return zip_filename
        else:
            log("ZIP file was not created", "ERROR")
            return None
            
    except Exception as e:
        log(f"Error creating ZIP archive: {e}", "ERROR")
        return None

def validate_email_address(email):
    """
    Basic email validation.
    
    Args:
        email (str): Email address to validate
    
    Returns:
        bool: True if email is valid, False otherwise
    """
    import re
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, email) is not None

def send_email_with_attachment(email_config, zip_file_path, month_year, total_amount, trip_count):
    """
    Send email with ZIP file attachment using Red Mail.
    
    Args:
        email_config (dict): Email configuration from config.json
        zip_file_path (str): Path to the ZIP file to attach
        month_year (str): Month-year string for email subject/body
        total_amount (float): Total trip amount for email body
        trip_count (int): Number of trips for email body
    
    Returns:
        bool: True if email sent successfully, False otherwise
    """
    log("Preparing to send email with ZIP attachment using Red Mail", "INFO")
    
    # Check if Red Mail is available
    if EmailSender is None:
        log("Red Mail library not installed. Please install it with: pip install redmail", "ERROR")
        return False
    
    # Validate email configuration
    required_fields = ['recipient_email', 'sender_email', 'sender_password', 'smtp_server', 'smtp_port']
    for field in required_fields:
        if field not in email_config or not email_config[field]:
            log(f"Missing required email configuration: {field}", "ERROR")
            return False
    
    # Validate email addresses
    if not validate_email_address(email_config['recipient_email']):
        log(f"Invalid recipient email: {email_config['recipient_email']}", "ERROR")
        return False
    
    if not validate_email_address(email_config['sender_email']):
        log(f"Invalid sender email: {email_config['sender_email']}", "ERROR")
        return False
    
    # Check if ZIP file exists
    if not os.path.exists(zip_file_path):
        log(f"ZIP file not found: {zip_file_path}", "ERROR")
        return False
    
    try:
        # Initialize Red Mail EmailSender
        log(f"Connecting to SMTP server: {email_config['smtp_server']}:{email_config['smtp_port']}", "INFO")
        
        email_sender = EmailSender(
            host=email_config['smtp_server'],
            port=email_config['smtp_port'],
            username=email_config['sender_email'],
            password=email_config['sender_password'],
            use_starttls=True  # Use STARTTLS for security
        )
        
        # Format subject
        subject_template = email_config.get('subject_template', 'Uber Trip Report - {month_year}')
        subject = subject_template.format(month_year=month_year)
        
        # Format body
        body_template = email_config.get('body_template', 
            'Please find attached the Uber trip report for {month_year}.\n\n'
            'Total amount: ${total_amount}\n'
            'Number of trips: {trip_count}\n\n'
            'Best regards,\n'
            'Uber Trip Exporter')
        
        body = body_template.format(
            month_year=month_year,
            total_amount=total_amount,
            trip_count=trip_count
        )
        
        # HTML version of the body for better formatting
        html_body = f"""
        <html>
        <body>
            <h2>Uber Trip Report - {month_year}</h2>
            <p>Please find attached the Uber trip report for <strong>{month_year}</strong>.</p>
            <ul>
                <li><strong>Total amount:</strong> ${total_amount}</li>
                <li><strong>Number of trips:</strong> {trip_count}</li>
            </ul>
            <p>Best regards,<br>Uber Trip Exporter</p>
        </body>
        </html>
        """
        
        log(f"Sending email to: {email_config['recipient_email']}", "INFO")
        
        # Send email with attachment using Red Mail
        email_sender.send(
            subject=subject,
            sender=email_config['sender_email'],
            receivers=[email_config['recipient_email']],
            text=body,
            html=html_body,
            attachments={
                os.path.basename(zip_file_path): Path(zip_file_path)
            }
        )
        
        log("Email sent successfully with Red Mail! 📧", "SUCCESS")
        return True
        
    except Exception as e:
        log(f"Error sending email with Red Mail: {e}", "ERROR")
        
        # Provide helpful error messages for common issues
        if "authentication" in str(e).lower():
            log("Authentication failed. Please check your email credentials.", "WARNING")
            log("For Gmail: Ensure you're using an App Password, not your regular password", "WARNING")
        elif "connection" in str(e).lower():
            log("Connection failed. Please check your SMTP server settings.", "WARNING")
        elif "permission" in str(e).lower() or "denied" in str(e).lower():
            log("Permission denied. Check your email provider's security settings.", "WARNING")
        
        return False    

# ============================================================================
# MAIN EXECUTION FUNCTION
# ============================================================================

def main():
    """Main function that orchestrates the entire Uber trip export process."""
    log("🚀 Starting Uber Trip Exporter", "HEADER")
    
    # Parse command line arguments
    target_month = parse_command_line_args()
    
    # Get date range for the target month
    start_time_ms, end_time_ms, month_year = get_month_date_range(target_month)
    
    # Read authentication token
    cookie = read_token_from_file()
    
    # Fetch trips from Uber API (with receipts using improved timeout handling)
    trips, overall_amount = get_uber_trips(cookie, start_time_ms, end_time_ms, download_receipts=True)
    
    if not trips:
        log("No trips found for the specified period", "WARNING")
        return
    
    # Create month-specific output folder
    output_folder = month_year
    os.makedirs(output_folder, exist_ok=True)
    log(f"Created output folder: {output_folder}", "SUCCESS")

    # Create month-specific output file paths
    monthly_receipts_file = os.path.join(output_folder, "all_receipts.pdf")
    monthly_trips_file = os.path.join(output_folder, "trips.json")

    # Save trips data with month prefix
    log("Saving trip data to JSON file", "INFO")
    with open(monthly_trips_file, "w", encoding="utf-8") as f:
        json.dump({
            "overall_amount": round(overall_amount, 2),
            "trips": trips,
            "month_year": month_year,
            "date_range": {
                "start": start_time_ms,
                "end": end_time_ms
            }
        }, f, ensure_ascii=False, indent=2)

    log(f"Saved {len(trips)} trips to {monthly_trips_file} (total: ${overall_amount:.2f})", "SUCCESS")
    
    # Merge receipts with month-specific filename
    merge_receipts(trips, output_file=monthly_receipts_file)
    
    # Clean up temporary receipts folder
    cleanup_temp_receipts_folder()

    # Load address keywords from config file
    log("Loading address configuration", "INFO")
    home_address_keywords, work_address_keywords, email_config = read_config_from_file()

    # Load the Excel template (original file name)
    template_excel_file = "Private_Taxi_Claim_Form.xlsx"

    # Create a monthly copy of the Excel file using the calculated month_year
    excel_file = create_monthly_excel_copy(template_excel_file, month_year, output_folder)

    # Process Excel file
    try:
        process_excel_file(excel_file, trips, template_excel_file, home_address_keywords, work_address_keywords)
        
        log("Monthly report completed successfully!", "HEADER")
        log(f"All files saved in folder: {output_folder}", "SUCCESS")
        log("Files created:", "INFO")
        log("  - trips.json (trip data)", "INFO")
        log("  - all_receipts.pdf (merged receipts)", "INFO")
        log(f"  - {month_year}_Private_Taxi_Claim_Form.xlsx (claim form)", "INFO")
        
        # Create ZIP archive and send email if enabled
        if email_config and email_config.get('enabled', False):
            log("Creating ZIP archive for email...", "INFO")
            zip_filename = f"{month_year}_uber_trip_report.zip"
            zip_path = create_zip_archive(output_folder, zip_filename)
            
            if zip_path:
                log("Sending email with ZIP attachment...", "INFO")
                email_sent = send_email_with_attachment(
                    email_config, 
                    zip_path, 
                    month_year, 
                    overall_amount, 
                    len(trips)
                )
                
                if email_sent:
                    log("Email sent successfully! 📧", "SUCCESS")
                    
                    # Optionally clean up the ZIP file after sending
                    try:
                        os.remove(zip_path)
                        log(f"Cleaned up ZIP file: {zip_path}", "INFO")
                    except Exception as e:
                        log(f"Could not clean up ZIP file: {e}", "WARNING")
                else:
                    log("Failed to send email. ZIP file preserved for manual sending.", "WARNING")
                    log(f"ZIP file location: {zip_path}", "INFO")
            else:
                log("Failed to create ZIP archive. Email not sent.", "ERROR")
        else:
            if email_config and not email_config.get('enabled', False):
                log("Email functionality is disabled in config. Set 'enabled' to true to send emails.", "INFO")
            else:
                log("No email configuration found. Files saved locally only.", "INFO")
        
    except Exception as e:
        log(f"Error processing Excel file: {e}", "ERROR")
        raise


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        log("Script interrupted by user", "WARNING")
        sys.exit(1)
    except Exception as e:
        log(f"Script failed with error: {e}", "ERROR")
        sys.exit(1)